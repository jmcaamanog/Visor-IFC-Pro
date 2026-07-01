# -*- coding: utf-8 -*-
"""
      ██╗███╗   ███╗ ██████╗
      ██║████╗ ████║██╔════╝
      ██║██╔████╔██║██║
 ██   ██║██║╚██╔╝██║██║    
 ╚█████╔╝██║ ╚═╝ ██║╚██████╗
  ╚════╝ ╚═╝     ╚═╝ ╚═════╝

---------------------------------------------------------------------------
 VISOR IFC v28.0 - Corrección Final de Estabilidad by JMC
---------------------------------------------------------------------------
 Descripción:
 Versión final completo y mejora la gestión de recursos de las ventanas.

 Mejoras Clave en v28.0:
  - CORRECCIÓN ESTABILIDAD 3D: Solucionado el cierre abrupto al interactuar
    con la aplicación después de cerrar la ventana del visor 3D completo.
    Se asegura que la referencia al plotter se limpie correctamente.
  - Mantenidas todas las funcionalidades y correcciones anteriores.
---------------------------------------------------------------------------
"""
import tkinter as tk
from tkinter import font, filedialog
import sys
import os
import traceback
import numpy as np
import threading
import webbrowser
import csv
from collections import Counter
from datetime import datetime
import locale
import io
import copy

# --- Importaciones de librerías con verificación ---
try:
    import ifcopenshell
    import ifcopenshell.api
    import ifcopenshell.util.element
    import ifcopenshell.geom
    IFCOPENDHELL_AVAILABLE = True
except ImportError:
    IFCOPENDHELL_AVAILABLE = False

try:
    import vedo
    VEDO_AVAILABLE = True
except ImportError:
    VEDO_AVAILABLE = False

try:
    from matplotlib.figure import Figure
    from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
    import matplotlib.pyplot as plt
    from matplotlib.ticker import MaxNLocator
    from matplotlib.patches import ConnectionPatch
    MATPLOTLIB_AVAILABLE = True
except ImportError:
    MATPLOTLIB_AVAILABLE = False
    
try:
    import ttkbootstrap as ttk
    from ttkbootstrap.dialogs import Messagebox
    from ttkbootstrap.scrolled import ScrolledFrame
    TTKBOOTSTRAP_AVAILABLE = True
except ImportError:
    TTKBOOTSTRAP_AVAILABLE = False
    
try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.drawing.image import Image as OpenpyxlImage
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    from PIL import Image, ImageTk
    PILLOW_AVAILABLE = True
except ImportError:
    PILLOW_AVAILABLE = False


# --- Clase Principal de la Aplicación ---
class IFCViewerApp:
    """
    Clase principal que encapsula toda la funcionalidad del visor IFC.
    """
    # --- Constantes de configuración ---
    COST_PARAM_NAMES = ["Costo", "Cost", "Coste"]
    AREA_PARAM_NAMES = ["Área", "Area", "NetFloorArea"]
    VOLUME_PARAM_NAMES = ["Volumen", "Volume", "NetVolume"]
    SPACE_NUMBER_PARAM_NAMES = ["Número"]
    SPACE_NAME_PARAM_NAMES = ["Nombre"]
    LEVEL_PARAM_NAMES = ["Nivel"]
    CLEAR_HEIGHT_PARAM_NAMES = ["Altura sin límites"]
    PHASE_PARAM_NAMES = ["FASE"]
    ASSEMBLY_CODE_PARAM_NAMES = ["Código de montaje", "Assembly Code"]
    KEYNOTE_PARAM_NAMES = ["Nota clave", "Keynote"]
    DESCRIPTION_PARAM_NAMES = ["Descripción", "Description"]
    TYPE_MARK_PARAM_NAMES = ["Marca de tipo", "Type Mark"]
    MARK_PARAM_NAMES = ["Marca", "Mark"]
    
    CONSTRUCTIVE_CLASSES = [
        "IfcWall", "IfcWallStandardCase", "IfcDoor", "IfcWindow", "IfcSlab",
        "IfcRoof", "IfcStair", "IfcStairFlight", "IfcRailing", "IfcBeam",
        "IfcColumn", "IfcFooting", "IfcPlate", "IfcMember", "IfcCovering",
        "IfcBuildingElementProxy"
    ]
    TRANSLATION_MAP = {
        "IfcWall": "Muros", "IfcWallStandardCase": "Muros", "IfcDoor": "Puertas",
        "IfcWindow": "Ventanas", "IfcSlab": "Forjados / Losas", "IfcStair": "Escaleras",
        "IfcStairFlight": "Tramos de Escalera", "IfcRailing": "Barandillas", "IfcBeam": "Vigas",
        "IfcColumn": "Pilares", "IfcRoof": "Cubiertas", "IfcFooting": "Cimentaciones", "IfcPlate": "Placas",
        "IfcBuildingElementProxy": "Elementos Constructivos", "IfcSpace": "Espacios",
        "IfcCovering": "Revestimientos", "IfcFlowTerminal": "Terminales de Fluido",
        "IfcPipeSegment": "Tramos de Tubería", "IfcMember": "Miembros / Montantes",
        "IfcFurnishingElement": "Mobiliario", "IfcGrid": "Ejes / Rejillas", "IfcOpeningElement": "Huecos"
    }
    COLOR_MAP_3D = {
        "IfcWall": "lightgrey", "IfcWallStandardCase": "lightgrey", "IfcDoor": "saddlebrown",
        "IfcWindow": "lightblue", "IfcSlab": "dimgrey", "IfcRoof": "darkred",
        "IfcStair": "orange", "IfcStairFlight": "orange", "IfcRailing": "grey",
        "IfcBeam": "lightcoral", "IfcColumn": "lightcoral", "IfcFooting": "darkgreen",
        "IfcFurnishingElement": "purple", "IfcBuildingElementProxy": "yellow"
    }

    def __init__(self, root):
        """Inicializa la aplicación, configura el estilo y crea la interfaz."""
        self.root = root
        self.root.title("Visor y Editor IFC by JMC")
        self.root.geometry("1707x1157")

        try:
            base_path = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
            icon_path = os.path.join(base_path, 'app_icon.ico')
            print(f"DEBUG: Buscando icono en: {icon_path}")
            if os.path.exists(icon_path):
                print("DEBUG: Icono encontrado.")
                self.root.iconbitmap(icon_path)
            else:
                print("DEBUG: Icono NO encontrado.")
        except Exception as e:
            print(f"Error al cargar el icono: {e}")
            traceback.print_exc()

        try:
            locale.setlocale(locale.LC_ALL, 'es_ES.UTF-8')
        except locale.Error:
            try:
                locale.setlocale(locale.LC_ALL, 'Spanish_Spain.1252')
            except locale.Error:
                locale.setlocale(locale.LC_ALL, '')

        self.style = ttk.Style.get_instance()
        self.base_font = ("Helvetica", 10)
        self.bold_font = ("Helvetica", 10, "bold")
        
        self.style.configure('.', font=self.base_font)
        self.style.configure('Treeview', rowheight=30)
        self.style.configure('Treeview.Heading', font=self.bold_font)
        self.style.configure('TButton', font=self.base_font)
        self.style.configure('Label', font=self.base_font)
        self.style.configure('group.TLabel', font=self.bold_font)
        self.header_bg = self.style.colors.inputbg
        self.style.configure('Metric.TLabel', font=('Helvetica', 14, 'bold'))
        self.style.configure('MetricTitle.TLabel', font=('Helvetica', 9), foreground=self.style.colors.secondary)
        self.style.configure('Total.Treeview', background=self.style.colors.light)
        self.style.configure('primary.TLabelframe.Label', font=self.bold_font)
        self.style.configure('info.TLabelframe.Label', font=self.bold_font)
        self.style.configure('TNotebook.Tab', font=self.bold_font)
        
        info_blue_color = self.style.colors.info 
        self.style.configure('PropName.TLabel', font=('Helvetica', 11, 'bold'), foreground=info_blue_color)
        self.style.configure('TotalCost.TLabel', font=('Helvetica', 20, 'bold'), foreground=info_blue_color)
        self.style.configure('Spaces.Treeview.Heading', font=('Helvetica', 11, 'bold'), background=self.style.colors.primary, foreground='white')

        self.ifc_file = None
        self.ifc_filepath = None
        self.tree_to_element_map = {}
        self.node_parent_map = {} 
        self.is_modified = False
        self.searchable_parameters = []
        
        self.vedo_plotter = None
        self.vedo_actors = {} 
        
        self.figures = {}
        self.logo_image = None
        
        self.total_project_area = 0
        self.total_project_volume = 0

        self.settings = ifcopenshell.geom.settings()
        self.settings.set(self.settings.USE_WORLD_COORDS, True)
        
        self.transparency_var = tk.BooleanVar(value=False)

        self._create_ui()
        
    def _create_ui(self):
        self._create_menubar()
        
        main_container = ttk.Frame(self.root, padding=(10, 10))
        main_container.pack(fill=tk.BOTH, expand=True)

        self._create_top_bar(main_container)

        main_pane = ttk.PanedWindow(main_container, orient=tk.HORIZONTAL)
        main_pane.pack(fill=tk.BOTH, expand=True)
        
        left_pane = ttk.Frame(main_pane, padding=5)
        main_pane.add(left_pane, weight=1)
        right_pane = ttk.Frame(main_pane, padding=5)
        main_pane.add(right_pane, weight=2)

        self._create_left_pane_widgets(left_pane)
        self._create_right_pane_widgets(right_pane)
        self._create_status_bar(main_container)

    def _create_menubar(self):
        menubar = tk.Menu(self.root)
        self.root.config(menu=menubar)

        self.file_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Archivo", menu=self.file_menu)
        self.file_menu.add_command(label="Abrir IFC...", command=self.select_file)
        self.file_menu.add_command(label="Guardar como...", command=self.save_modified_ifc, state="disabled")
        
        export_menu = tk.Menu(self.file_menu, tearoff=0)
        self.file_menu.add_cascade(label="Exportar...", menu=export_menu)
        export_menu.add_command(label="Informe de Modelo (Excel)", command=lambda: self.open_export_dialog(mode='model'))
        export_menu.add_command(label="Elemento Seleccionado (Excel)", command=lambda: self.open_export_dialog(mode='element'))
        export_menu.add_separator()
        export_menu.add_command(label="Resumen del Modelo (CSV)", command=self.export_summary_to_csv)
        export_menu.add_command(label="Propiedades de Elemento (CSV)", command=self.export_properties_to_csv)
        
        self.file_menu.add_separator()
        self.file_menu.add_command(label="Salir", command=self.root.destroy)

        self.view_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Ver", menu=self.view_menu)
        self.view_menu.add_command(label="Ver Modelo 3D Completo", command=self.start_full_model_view)
        self.view_menu.add_separator()
        self.view_menu.add_checkbutton(label="Modo Transparencia", variable=self.transparency_var, command=self.toggle_transparency_mode)
        
        help_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Ayuda", menu=help_menu)
        help_menu.add_command(label="Acerca de...", command=lambda: self.notebook.select(5))

    def _create_top_bar(self, parent):
        top_bar = ttk.Frame(parent)
        top_bar.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Button(top_bar, text="Abrir IFC...", command=self.select_file, bootstyle="primary").pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(top_bar, text="Ver Modelo 3D", command=self.start_full_model_view, bootstyle="secondary").pack(side=tk.LEFT, padx=(0, 5))
        
        self.view_selected_button = ttk.Button(top_bar, text="Ver Elemento 3D", command=self.view_selected_element_3d, bootstyle="info", state="disabled")
        self.view_selected_button.pack(side=tk.LEFT, padx=(0, 10))
        
        self.file_path_label = ttk.Label(top_bar, text="Archivo: Ninguno seleccionado", anchor="w")
        self.file_path_label.pack(fill=tk.X, expand=True, side=tk.LEFT)

    def _create_left_pane_widgets(self, parent):
        search_frame = ttk.Frame(parent)
        search_frame.pack(fill=tk.X, pady=(0, 5))
        self.search_entry = ttk.Entry(search_frame)
        self.search_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0,5))
        self.search_entry.bind("<Return>", self.simple_search)

        ttk.Button(search_frame, text="Buscar", command=self.simple_search, bootstyle="info").pack(side=tk.LEFT, padx=(0,2))
        ttk.Button(search_frame, text="Limpiar", command=self.clear_simple_search, bootstyle="outline-secondary").pack(side=tk.LEFT, padx=(0,5))
        ttk.Button(search_frame, text="Avanzada...", command=self.open_advanced_search, bootstyle="outline-info").pack(side=tk.LEFT)

        tree_container = ttk.Labelframe(parent, text="ESTRUCTURA DEL PROYECTO", bootstyle="info")
        tree_container.pack(fill=tk.BOTH, expand=True, pady=(5,0))
        self.project_tree = ttk.Treeview(tree_container, bootstyle="info")
        self.project_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        tree_scrollbar = ttk.Scrollbar(tree_container, orient="vertical", command=self.project_tree.yview, bootstyle="info-round")
        tree_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.project_tree.config(yscrollcommand=tree_scrollbar.set)
        self.project_tree.bind('<<TreeviewSelect>>', self.on_tree_select)

    def _create_right_pane_widgets(self, parent):
        self.key_props_labels = {}
        self.marks_labels = {}

        key_props_frame = ttk.Labelframe(parent, text="PROPIEDADES CLAVE", bootstyle="info", padding=10)
        key_props_frame.pack(fill=tk.X, pady=(0, 10))
        key_props_map = ["CATEGORÍA", "FAMILIA", "TIPO", "NIVEL", "ÁREA", "VOLUMEN"]
        for i, prop_name in enumerate(key_props_map):
            row, col = divmod(i, 2)
            ttk.Label(key_props_frame, text=f"{prop_name}:", style='PropName.TLabel').grid(row=row, column=col*2, sticky="w", padx=(0,5))
            value_label = ttk.Label(key_props_frame, text="N/A")
            value_label.grid(row=row, column=col*2+1, sticky="w")
            key_props_frame.grid_columnconfigure(col*2+1, weight=1)
            self.key_props_labels[prop_name] = value_label

        marks_frame = ttk.Labelframe(parent, text="MARCAS", bootstyle="info", padding=10)
        marks_frame.pack(fill=tk.X, pady=(0, 10))
        marks_map = ["CÓDIGO DE MONTAJE", "NOTA CLAVE", "MARCA DE TIPO", "MARCA", "ID DE TIPO", "GLOBALID"]
        for i, prop_name in enumerate(marks_map):
            row, col = divmod(i, 2)
            ttk.Label(marks_frame, text=f"{prop_name}:", style='PropName.TLabel').grid(row=row, column=col*2, sticky="w", padx=(0,5))
            value_label = ttk.Label(marks_frame, text="N/A", wraplength=300)
            value_label.grid(row=row, column=col*2+1, sticky="w")
            marks_frame.grid_columnconfigure(col*2+1, weight=1)
            self.marks_labels[prop_name] = value_label
            
        desc_frame = ttk.Labelframe(parent, text="DESCRIPCIÓN", bootstyle="info", padding=10)
        desc_frame.pack(fill=tk.X, pady=(0, 10))
        self.description_text_label = ttk.Label(desc_frame, text="N/A", justify=tk.LEFT)
        self.description_text_label.pack(fill=tk.X, expand=True)
        self.description_text_label.bind('<Configure>', lambda e: e.widget.config(wraplength=e.widget.winfo_width()))
        
        self.notebook = ttk.Notebook(parent, bootstyle="dark")
        self.notebook.pack(fill=tk.BOTH, expand=True, pady=(5,0))
        self._create_tabs()

        bottom_buttons_frame = ttk.Frame(parent)
        bottom_buttons_frame.pack(fill=tk.X, pady=(10, 0))
        self.export_button = ttk.Button(bottom_buttons_frame, text="Exportar Info (Excel)", command=lambda: self.open_export_dialog(mode='model'), state="disabled", bootstyle="info-outline")
        self.export_button.pack(side=tk.RIGHT, padx=(5,0))
        self.save_button = ttk.Button(bottom_buttons_frame, text="Guardar IFC Modificado", command=self.save_modified_ifc, state="disabled", bootstyle="success-outline")
        self.save_button.pack(side=tk.RIGHT)

    def _create_tabs(self):
        props_tab = ttk.Frame(self.notebook)
        self.notebook.add(props_tab, text='PROPIEDADES ELEMENTO')
        self.properties_tree = ttk.Treeview(props_tab, columns=('Value'), bootstyle="info")
        self.properties_tree.heading('#0', text='Propiedad'); self.properties_tree.heading('Value', text='Valor')
        self.properties_tree.column('#0', width=250, stretch=tk.YES); self.properties_tree.column('Value', width=400, stretch=tk.YES)
        self.properties_tree.bind('<Double-1>', self.on_property_double_click)
        self.properties_tree.pack(fill=tk.BOTH, expand=True)
        self.properties_tree.tag_configure('edited', foreground=self.style.colors.info, font=self.bold_font)
        self.properties_tree.tag_configure('group_header', font=self.bold_font, background=self.header_bg)

        self._create_cost_tab()
        self._create_summary_tab()
        self._create_spaces_tab()
        self._create_analytics_tab()
        self._create_about_tab()

    def _create_cost_tab(self):
        self.cost_tab = ttk.Frame(self.notebook, padding=10)
        self.notebook.add(self.cost_tab, text='COSTE')

        canvas = tk.Canvas(self.cost_tab, bg=self.style.colors.bg, highlightthickness=0)
        scrollbar = ttk.Scrollbar(self.cost_tab, orient="vertical", command=canvas.yview, bootstyle="info-round")
        self.cost_content_frame = ttk.Frame(canvas, bootstyle="dark")
        self.cost_content_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=self.cost_content_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        top_metrics_frame = ttk.Frame(self.cost_content_frame)
        top_metrics_frame.pack(fill='x', expand=True, padx=10, pady=10)
        top_metrics_frame.columnconfigure(0, weight=1)
        top_metrics_frame.columnconfigure(1, weight=2)

        economic_frame = ttk.Labelframe(top_metrics_frame, text="RESUMEN ECONÓMICO", bootstyle="primary", padding=15)
        economic_frame.grid(row=0, column=0, sticky="nsew", padx=(0,5))
        economic_frame.columnconfigure(0, weight=1)
        self.metric_labels = {}
        self.metric_labels['cost'] = self._create_metric_widget(economic_frame, "Coste Total", "N/A", style="TotalCost.TLabel", bootstyle="info")
        self.metric_labels['cost_sqm'] = self._create_metric_widget(economic_frame, "Coste / m²", "N/A")

        geometry_frame = ttk.Labelframe(top_metrics_frame, text="GEOMETRÍA DEL PROYECTO", bootstyle="info", padding=15)
        geometry_frame.grid(row=0, column=1, sticky="nsew", padx=(5,0))
        geometry_frame.columnconfigure((0,1), weight=1)
        self.metric_labels['area'] = self._create_metric_widget(geometry_frame, "Superficie Total Útil", "N/A", "info", 0, 0)
        self.metric_labels['volume'] = self._create_metric_widget(geometry_frame, "Volumen Total", "N/A", "info", 1, 0)
        self.metric_labels['spaces'] = self._create_metric_widget(geometry_frame, "Nº de Habitaciones", "N/A", "info", 0, 1)
        self.metric_labels['levels'] = self._create_metric_widget(geometry_frame, "Nº de Niveles", "N/A", "info", 1, 1)

        self.cost_bar_chart_frame = ttk.Labelframe(self.cost_content_frame, text="DESGLOSE DE COSTE POR TIPO", bootstyle="info", padding=10)
        self.cost_bar_chart_frame.pack(fill="x", padx=10, pady=10, expand=True)
        
        self.timestamp_label = ttk.Label(self.cost_content_frame, text="El coste es una estimación basada en los datos del modelo.", font=("Helvetica", 11, "italic"), bootstyle="info", anchor="center")
        self.timestamp_label.pack(fill="x", pady=(20, 10), padx=10)

    def _create_summary_tab(self):
        summary_container_frame = ttk.Frame(self.notebook, padding=0)
        self.notebook.add(summary_container_frame, text='RESUMEN')
        self.summary_tab = ScrolledFrame(summary_container_frame, autohide=True, bootstyle="dark-round")
        self.summary_tab.pack(fill=tk.BOTH, expand=tk.YES)

    def _create_spaces_tab(self):
        self.spaces_tab = ttk.Frame(self.notebook, padding=(5,10))
        self.notebook.add(self.spaces_tab, text='ESPACIOS')
        
        project_info_frame = ttk.Labelframe(self.spaces_tab, text="LISTADO DE ESPACIOS", bootstyle="info", padding=5)
        project_info_frame.pack(fill="both", expand=True)
        
        self.spaces_tree = ttk.Treeview(project_info_frame, columns=('Name', 'Level', 'ClearHeight', 'Phase', 'Area', 'Volume'), bootstyle="info", style="Spaces.Treeview")
        self.spaces_tree.heading('#0', text='Número') 
        self.spaces_tree.heading('Name', text='Nombre')
        self.spaces_tree.heading('Level', text='Nivel')
        self.spaces_tree.heading('ClearHeight', text='Altura Libre (m)')
        self.spaces_tree.heading('Phase', text='Fase')
        self.spaces_tree.heading('Area', text='Superficie (m²)')
        self.spaces_tree.heading('Volume', text='Volumen (m³)')
        self.spaces_tree.column('#0', width=80, anchor='center', stretch=tk.NO)
        self.spaces_tree.column('Name', stretch=tk.YES)
        self.spaces_tree.column('Level', width=120, anchor='center', stretch=tk.NO)
        self.spaces_tree.column('ClearHeight', width=120, anchor='center', stretch=tk.NO)
        self.spaces_tree.column('Phase', width=100, anchor='center', stretch=tk.NO)
        self.spaces_tree.column('Area', width=120, anchor='center', stretch=tk.NO)
        self.spaces_tree.column('Volume', width=120, anchor='center', stretch=tk.NO)
        self.spaces_tree.pack(fill="both", expand=True)
        self.spaces_tree.tag_configure('total_row', font=('Helvetica', 11, 'bold'), background=self.style.colors.light)
        self.spaces_tree.tag_configure('spacer_row', background=self.style.colors.bg)

    def _create_metric_widget(self, parent, title, default_val, bootstyle="primary", grid_row=None, grid_col=None, style='Metric.TLabel'):
        container = ttk.Frame(parent)
        if grid_row is not None:
            container.grid(row=grid_row, column=grid_col, sticky="nsew", padx=5, pady=5)
        else:
            container.pack(fill='x', pady=5)
            
        title_label = ttk.Label(container, text=title, style='MetricTitle.TLabel', anchor="center")
        title_label.pack(fill='x')
        value_label = ttk.Label(container, text=default_val, style=style, anchor="center", bootstyle=bootstyle)
        value_label.pack(fill='x')
        return value_label

    def _create_status_bar(self, parent):
        status_frame = ttk.Frame(parent, padding=(5, 2))
        status_frame.pack(fill=tk.X, side=tk.BOTTOM, pady=(10,0))
        self.progress_label = ttk.Label(status_frame, text="Listo.")
        self.progress_label.pack(side=tk.LEFT, padx=5)
        self.progress_bar = ttk.Progressbar(status_frame, mode='determinate', bootstyle="info")
        self.progress_bar.pack(side=tk.RIGHT, padx=5, fill=tk.X, expand=True)

    def _create_analytics_tab(self):
        self.analytics_tab = ttk.Frame(self.notebook, padding=10)
        self.notebook.add(self.analytics_tab, text='ANÁLISIS')
        
        canvas = tk.Canvas(self.analytics_tab, bg=self.style.colors.bg, highlightthickness=0)
        scrollbar = ttk.Scrollbar(self.analytics_tab, orient="vertical", command=canvas.yview, bootstyle="info-round")
        self.analytics_content_frame = ttk.Frame(canvas, bootstyle="dark")
        self.analytics_content_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=self.analytics_content_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

    def _create_about_tab(self):
        about_tab = ttk.Frame(self.notebook, padding=40)
        self.notebook.add(about_tab, text='Acerca de...')
        
        center_frame = ttk.Frame(about_tab)
        center_frame.pack(expand=True)
        
        linkedin_url = "https://www.linkedin.com/in/josecaamano/"
        open_linkedin = lambda e: webbrowser.open_new(linkedin_url)

        try:
            base_path = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
            logo_path = os.path.join(base_path, 'logo.png')
            print(f"DEBUG: Buscando logo en: {logo_path}")
            if os.path.exists(logo_path):
                print("DEBUG: Logo encontrado.")
                pil_image = Image.open(logo_path)
                pil_image = pil_image.resize((128, 128), Image.Resampling.LANCZOS)
                self.logo_image = ImageTk.PhotoImage(pil_image)
                logo_label = ttk.Label(center_frame, image=self.logo_image, cursor="hand2")
                logo_label.pack(pady=(0, 15))
                logo_label.bind("<Button-1>", open_linkedin)
            else:
                print("DEBUG: Logo NO encontrado.")
        except Exception as e:
            print(f"No se pudo cargar el logo: {e}")

        ttk.Label(center_frame, text="Visor IFC Pro", font=("Helvetica", 24, "bold"), bootstyle="primary").pack(pady=(0, 5))
        ttk.Label(center_frame, text="De un BIM Manager, para todo el mundo.", font=("Helvetica", 12), bootstyle="secondary").pack()
        
        ttk.Separator(center_frame, bootstyle="secondary").pack(fill='x', pady=20, padx=50)
        info_text = ("Hecho con código y café desde A Coruña. ☕\n\n"
                     "Esta aplicación nace de la trinchera, de la mente de un\n"
                     "Arquitecto Técnico y BIM Manager que cree que las\n"
                     "herramientas deben ser potentes, pero también sencillas.\n\n"
                     "Aquí no hay adornos innecesarios, solo la funcionalidad que\n"
                     "necesitas para visualizar, analizar y editar tus modelos IFC.")
        ttk.Label(center_frame, text=info_text, justify=tk.CENTER, font=("Helvetica", 11)).pack(pady=10)
        ttk.Separator(center_frame, bootstyle="secondary").pack(fill='x', pady=20, padx=50)
        ttk.Label(center_frame, text="Jose Manuel Caamaño González", font=("Helvetica", 12, "bold")).pack()
        link_font = font.Font(family="Helvetica", size=11, underline=True)
        link_label = tk.Label(center_frame, text="josecaamano.io", fg=self.style.colors.info, bg=self.style.colors.bg, cursor="hand2", font=link_font)
        link_label.pack()
        link_label.bind("<Button-1>", lambda e: webbrowser.open_new("http://josecaamano.io"))
        
        jmc_label = ttk.Label(center_frame, text="BY JMC", font=("Courier", 16, "bold"), bootstyle="secondary", cursor="hand2")
        jmc_label.pack(pady=(20, 0))
        jmc_label.bind("<Button-1>", open_linkedin)

        ttk.Label(center_frame, text="Copyright 2025", font=("Helvetica", 9), bootstyle="secondary").pack(pady=(5, 0))

    def select_file(self):
        filepath = filedialog.askopenfilename(
            title="Seleccionar archivo IFC", filetypes=(("Archivos IFC", "*.ifc"), ("Todos los archivos", "*.*"))
        )
        if filepath:
            self.ifc_filepath = filepath
            self.file_path_label.config(text=f"Archivo: {os.path.basename(filepath)}")
            threading.Thread(target=self.load_and_process_ifc, args=(filepath,), daemon=True).start()

    def load_and_process_ifc(self, filepath):
        try:
            self.update_status("Cargando archivo IFC...", 0)
            self.ifc_file = ifcopenshell.open(filepath)
            
            self.root.after(0, self.clear_all_views)
            
            self.update_status("Procesando estructura...", 20)
            project_entity = self.ifc_file.by_type('IfcProject')[0]
            self.root.after(0, self.populate_project_tree, project_entity, '')
            
            self.update_status("Generando resumen...", 40)
            self.root.after(0, self.populate_summary_tab)

            self.update_status("Calculando métricas...", 50)
            self.root.after(0, self.update_spaces_tab)
            self.root.after(0, self.update_cost_tab)

            self.update_status("Creando análisis gráfico...", 60)
            self.root.after(0, self.update_analytics_tab)
            
            self.update_status("Indexando parámetros...", 80)
            self.collect_searchable_parameters()
            
            self.root.after(0, lambda: self.export_button.config(state="normal"))
            self.root.after(0, self.notebook.select, 0)
            self.update_status("Listo.", 100)
            self.root.after(1000, lambda: self.update_status("Listo.", 0))
            
            now = datetime.now()
            time_str = now.strftime("%d/%m/%Y | %H:%M")
            project_name = getattr(project_entity, 'Name', 'Sin Nombre')
            message = f"Proyecto '{project_name}' cargado con éxito.\n\n{time_str}\n(by JMC)"
            self.root.after(0, lambda: Messagebox.ok(message, "¡Carga Completa!"))
            
        except Exception:
            error_info = traceback.format_exc()
            self.root.after(0, lambda: handle_error(error_info, self.root))
            self.update_status("Error al cargar el archivo.", 0)
            
    def update_status(self, message, progress):
        def callback():
            self.progress_label.config(text=message)
            self.progress_bar.configure(value=progress)
        self.root.after(0, callback)

    def clear_all_views(self):
        self.save_button.config(state="disabled")
        self.export_button.config(state="disabled")
        self.file_menu.entryconfig("Guardar como...", state="disabled")
        self.is_modified = False
        
        if self.vedo_plotter:
            try:
                self.vedo_plotter.close()
            except: pass
            self.vedo_plotter = None
        self.vedo_actors.clear()
        
        self.project_tree.delete(*self.project_tree.get_children())
        self.properties_tree.delete(*self.properties_tree.get_children())
        
        for widget in self.summary_tab.winfo_children(): widget.destroy()
        if hasattr(self, 'cost_content_frame'):
            for label in self.metric_labels.values(): label.config(text="N/A")
            for widget in self.cost_bar_chart_frame.winfo_children(): widget.destroy()
        if hasattr(self, 'spaces_tree'):
            self.spaces_tree.delete(*self.spaces_tree.get_children())
        if hasattr(self, 'spaces_donut_chart_frame'):
            for widget in self.spaces_donut_chart_frame.winfo_children(): widget.destroy()
        if hasattr(self, 'analytics_content_frame'):
            for widget in self.analytics_content_frame.winfo_children(): widget.destroy()
            
        self.search_entry.delete(0, tk.END)
        for label in self.key_props_labels.values(): label.config(text="N/A")
        for label in self.marks_labels.values(): label.config(text="N/A")
        if hasattr(self, 'description_text_label'): self.description_text_label.config(text="N/A")

        self.tree_to_element_map.clear()
        self.node_parent_map.clear()
        self.figures.clear()
    
    def _get_property_from_psets(self, psets, param_names, pset_name=None):
        if not psets: return None
        
        if pset_name:
            target_pset = psets.get(pset_name)
            if target_pset:
                for name in param_names:
                    if name in target_pset:
                        return target_pset[name]
        else:
            for pset in psets.values():
                for name in param_names:
                    if name in pset:
                        return pset[name]
        return None

    def _create_mesh_from_shape(self, shape):
        try:
            verts = np.array(shape.geometry.verts).reshape((-1, 3))
            faces = np.array(shape.geometry.faces).reshape((-1, 3))
            if not (verts.size and faces.size):
                return None
            return vedo.Mesh([verts, faces])
        except (AttributeError, IndexError):
            return None

    def start_full_model_view(self):
        if not self.ifc_file: 
            Messagebox.show_warning("Por favor, carga primero un archivo IFC.", "Sin Archivo")
            return
        threading.Thread(target=self.display_full_model_popup, daemon=True).start()

    def display_full_model_popup(self):
        self.update_status("Cargando elementos 3D...", 0)
        
        if self.vedo_plotter and self.vedo_plotter.window:
             self.vedo_plotter.close()

        self.vedo_actors.clear()
        all_meshes = []
        products = self.ifc_file.by_type('IfcProduct')
        
        for i, product in enumerate(products):
            if not product.Representation: continue
            try:
                progress = (i + 1) / len(products) * 100
                self.update_status(f"Generando malla para {product.is_a()}...", progress)
                
                shape = ifcopenshell.geom.create_shape(self.settings, product)
                mesh = self._create_mesh_from_shape(shape)
                
                if mesh:
                    color = self.COLOR_MAP_3D.get(product.is_a(), "white")
                    actor = mesh.color(color).backface_culling(True)
                    actor.name = str(product.id())
                    self.vedo_actors[product.id()] = actor
                    all_meshes.append(actor)
            except Exception:
                continue
                
        if not all_meshes:
            self.root.after(0, lambda: Messagebox.show_warning("No se pudo generar ninguna geometría visible.", "Visor 3D"))
            self.update_status("Listo.", 0)
            return

        self.update_status("Abriendo visor 3D...", 100)
        
        self.vedo_plotter = vedo.Plotter(title="Visor 3D - Modelo Completo", axes=1)
        
        if self.transparency_var.get():
             self.toggle_transparency_mode()

        self.vedo_plotter.show(all_meshes, "Modelo Completo", bg='black', bg2='darkblue', interactive=True)
        
        # --- CORRECCIÓN CRÍTICA: Limpiar la referencia al plotter cuando se cierra la ventana ---
        self.vedo_plotter = None
        self.vedo_actors.clear()
        
        self.update_status("Listo.", 0)

    def view_selected_element_3d(self):
        selected_items = self.project_tree.selection()
        if not selected_items:
            Messagebox.show_warning("Por favor, selecciona un elemento en el árbol primero.", "Ningún Elemento Seleccionado")
            return
            
        element = self.tree_to_element_map.get(selected_items[0])
        if not element: return
        
        threading.Thread(target=self._create_and_show_single_element_viewer, args=(element,), daemon=True).start()

    def _create_and_show_single_element_viewer(self, element):
        """Genera y muestra un visor 3D para un único elemento de forma robusta."""
        plotter = None
        try:
            self.update_status(f"Generando vista para {element.Name or element.is_a()}...", 50)
            
            shape = ifcopenshell.geom.create_shape(self.settings, element)
            mesh = self._create_mesh_from_shape(shape)
            
            if not mesh:
                self.root.after(0, lambda: Messagebox.show_warning("No se pudo generar la geometría para este elemento.", "Error de Geometría"))
                self.update_status("Listo.", 0)
                return
            
            actor = mesh.color('lightblue').backface_culling(True)
            
            plotter = vedo.Plotter(title=f"Visor de Elemento: {element.Name or element.GlobalId}", axes=1)
            plotter.show(actor, bg='black', bg2='darkblue', interactive=True).close()
            
        except Exception:
            error_info = traceback.format_exc()
            self.root.after(0, lambda: handle_error(error_info, self.root))
        finally:
            if plotter:
                try:
                    plotter.close()
                except:
                    pass
            self.update_status("Listo.", 0)


    def _get_selected_element(self):
        selected_items = self.project_tree.selection()
        if not selected_items:
            return None
        return self.tree_to_element_map.get(selected_items[0])

    def highlight_element_in_full_model(self, element):
        if not self.vedo_plotter or not self.vedo_plotter.window or not self.vedo_actors: return
        
        for actor in self.vedo_actors.values():
            ifc_class = self.ifc_file.by_id(int(actor.name)).is_a()
            actor.color(self.COLOR_MAP_3D.get(ifc_class, "white")).alpha(1.0)

        actor_to_highlight = self.vedo_actors.get(element.id())
        if actor_to_highlight:
            actor_to_highlight.color("red")
        
        self.vedo_plotter.render()
        
    def toggle_transparency_mode(self):
        if not self.vedo_plotter or not self.vedo_plotter.window or not self.vedo_actors:
            return

        is_on = self.transparency_var.get()
        
        for actor in self.vedo_actors.values():
            actor.on().alpha(0.1 if is_on else 1.0)

        if is_on:
            selected_element = self._get_selected_element()
            if selected_element:
                actor_to_highlight = self.vedo_actors.get(selected_element.id())
                if actor_to_highlight:
                    actor_to_highlight.alpha(1.0).color("red")
        elif not is_on:
            selected_element = self._get_selected_element()
            if selected_element:
                self.highlight_element_in_full_model(selected_element)

        self.vedo_plotter.render()
    
    def update_analytics_tab(self):
        if not self.ifc_file: return
        for widget in self.analytics_content_frame.winfo_children(): widget.destroy()
        
        level_data_detailed = {}
        level_data_summary = {}
        constructive_elements_summary = Counter()
        
        for element in self.ifc_file.by_type("IfcProduct"):
            if element.is_a() not in self.CONSTRUCTIVE_CLASSES: continue
            
            container = ifcopenshell.util.element.get_container(element)
            level_name = container.Name if container and container.is_a("IfcBuildingStorey") else "Sin Nivel Asignado"
            type_name = self.TRANSLATION_MAP.get(element.is_a(), element.is_a())
            
            level_data_summary.setdefault(level_name, 0)
            level_data_summary[level_name] += 1
            
            level_data_detailed.setdefault(level_name, Counter())
            level_data_detailed[level_name][type_name] += 1

            constructive_elements_summary[type_name] += 1

        if not level_data_summary:
            ttk.Label(self.analytics_content_frame, text="No se encontraron elementos constructivos para analizar.").pack(expand=True, pady=20)
            return
        
        self._draw_summary_bar_chart(level_data_summary)
        ttk.Separator(self.analytics_content_frame).pack(fill='x', pady=20, padx=10)

        self._draw_elements_by_level_chart(level_data_detailed)
        ttk.Separator(self.analytics_content_frame).pack(fill='x', pady=20, padx=10)

        self._draw_donut_chart(constructive_elements_summary)

    def _draw_summary_bar_chart(self, summary_data):
        container = ttk.Labelframe(self.analytics_content_frame, text="RECUENTO TOTAL POR NIVEL", bootstyle="info", padding=10)
        container.pack(fill="x", expand=True, padx=10, pady=10)

        sorted_levels = sorted(summary_data.items(), key=lambda item: item[1], reverse=True)
        data = dict(reversed(list(dict(sorted_levels).items())))

        bg_color, primary_color = self.style.colors.bg, self.style.colors.primary
        text_color = 'white' 

        labels, values = list(data.keys()), list(data.values())
        
        fig = Figure(figsize=(10, len(labels) * 0.5), dpi=100, facecolor=bg_color)
        self.figures['summary'] = fig
        ax = fig.add_subplot(111)
        ax.set_facecolor(bg_color)
        
        colors = plt.cm.viridis(np.linspace(0.4, 0.9, len(labels)))
        y_pos = np.arange(len(labels))
        ax.barh(y_pos, values, color=colors, align='center', height=0.6)

        ax.spines['top'].set_visible(False); ax.spines['right'].set_visible(False)
        ax.spines['left'].set_visible(False); ax.spines['bottom'].set_color(text_color)
        ax.set_yticks([])
        ax.tick_params(axis='x', colors=text_color)
        ax.xaxis.set_major_locator(MaxNLocator(integer=True))

        for i, (label, value) in enumerate(data.items()):
            level_text = f" {label.upper()}"
            count_text = f" {value} Elementos"
            ax.text(-0.01 * max(values), y_pos[i], level_text, ha='right', va='center', color=primary_color, weight='bold', fontsize=9)
            ax.text(0.01 * max(values), y_pos[i], count_text, ha='left', va='center', color=text_color, weight='normal', fontsize=9)

        fig.tight_layout(pad=1.5)
        
        canvas = FigureCanvasTkAgg(fig, master=container)
        canvas.draw()
        canvas.get_tk_widget().pack(fill='both', expand=True, pady=5)

    def _draw_elements_by_level_chart(self, detailed_data):
        container = ttk.Labelframe(self.analytics_content_frame, text="ELEMENTOS POR NIVELES", bootstyle="info", padding=10)
        container.pack(fill="x", expand=True, padx=10, pady=10)

        df = pd.DataFrame(detailed_data).fillna(0).astype(int)
        
        all_types = sorted(df.index)
        cost_colors = plt.cm.viridis(np.linspace(0.1, 0.9, len(all_types)))
        color_map = {category: color for category, color in zip(all_types, cost_colors)}
        
        bg_color, text_color = self.style.colors.bg, self.style.colors.fg
        
        df_transposed = df.T
        
        df_transposed['Total'] = df_transposed.sum(axis=1)
        df_transposed = df_transposed.sort_values(by='Total', ascending=True)
        df_transposed = df_transposed.drop(columns=['Total'])

        fig = Figure(figsize=(10, len(df_transposed) * 0.6 + 1), dpi=100, facecolor=bg_color)
        self.figures['elements_by_level'] = fig
        ax = fig.add_subplot(111)
        ax.set_facecolor(bg_color)
        
        plot_colors = [color_map.get(col) for col in df_transposed.columns]
        
        df_transposed.plot(kind='barh', stacked=True, ax=ax, color=plot_colors, width=0.7)

        ax.spines['top'].set_visible(False); ax.spines['right'].set_visible(False)
        ax.spines['left'].set_color(text_color); ax.spines['bottom'].set_color(text_color)
        
        ax.tick_params(axis='x', colors=text_color)
        ax.tick_params(axis='y', colors=text_color, labelrotation=0)
        ax.set_xlabel('Cantidad de Elementos', color=text_color)
        ax.set_ylabel('')
        
        leg = ax.legend(title="Categorías", bbox_to_anchor=(1.01, 1), loc='upper left', labelcolor=text_color, frameon=False)
        leg.get_title().set_color(text_color)
        
        fig.tight_layout(rect=[0, 0.05, 0.85, 1])

        canvas = FigureCanvasTkAgg(fig, master=container)
        canvas.draw()
        canvas.get_tk_widget().pack(fill='both', expand=True, pady=5)


    def _draw_donut_chart(self, summary_data):
        container = ttk.Labelframe(self.analytics_content_frame, text="DISTRIBUCIÓN PORCENTUAL DE ELEMENTOS", bootstyle="info", padding=10)
        container.pack(fill="x", expand=True, padx=10, pady=10)

        total_elements = sum(summary_data.values())
        threshold = 0.03
        
        main_data = {}
        others_sum = 0
        others_count = 0
        
        sorted_data = dict(sorted(summary_data.items(), key=lambda item: item[1], reverse=True))

        for key, value in sorted_data.items():
            if (value / total_elements) < threshold:
                others_sum += value
                others_count += 1
            else:
                main_data[key] = value

        if others_sum > 0:
            main_data[f'Elementos Varios ({others_count})'] = others_sum

        data = dict(sorted(main_data.items(), key=lambda item: item[1], reverse=True))
        labels = data.keys()
        sizes = list(data.values())

        bg_color, text_color = self.style.colors.bg, self.style.colors.fg
        colors = plt.cm.viridis(np.linspace(0.1, 0.9, len(labels)))

        fig = Figure(figsize=(10, 6), dpi=100, facecolor=bg_color)
        self.figures['donut'] = fig
        ax = fig.add_subplot(111)
        ax.set_facecolor(bg_color)

        wedges, texts, autotexts = ax.pie(sizes, 
                                          autopct='%1.1f%%', 
                                          startangle=90, 
                                          colors=colors,
                                          pctdistance=0.8,
                                          wedgeprops=dict(width=0.4, edgecolor=bg_color))
        
        plt.setp(autotexts, size=9, weight="bold", color="white")
        ax.axis('equal')

        legend_labels = [f'{l} ({s})' for l, s in zip(labels, sizes)]
        leg = ax.legend(wedges, legend_labels,
                          title="Tipos de Elemento",
                          loc="center left",
                          bbox_to_anchor=(1.1, 0.5),
                          labelcolor=text_color,
                          frameon=False)
        leg.get_title().set_color(text_color)
        
        fig.tight_layout()
        canvas = FigureCanvasTkAgg(fig, master=container)
        canvas.draw()
        canvas.get_tk_widget().pack(fill='both', expand=True, pady=5)


    def on_property_double_click(self, event):
        region = self.properties_tree.identify("region", event.x, event.y)
        if region != "cell": return

        column = self.properties_tree.identify_column(event.x)
        if column != "#1": return

        item_id = self.properties_tree.identify_row(event.y)
        parent_id = self.properties_tree.parent(item_id)
        if not parent_id: return

        prop_name = self.properties_tree.item(item_id, 'text')
        parent_text = self.properties_tree.item(parent_id, 'text')

        is_editable_attribute = parent_text == "INFORMACIÓN GENERAL" and prop_name in ["Nombre", "Description"]
        grandparent_id = self.properties_tree.parent(parent_id)
        is_pset_property = grandparent_id and self.properties_tree.item(grandparent_id, 'text') == "PROPIEDADES (PROPERTY SETS)"
        
        if not (is_editable_attribute or is_pset_property): return

        x, y, width, height = self.properties_tree.bbox(item_id, column)
        value = self.properties_tree.set(item_id, column)

        entry = ttk.Entry(self.properties_tree, justify='left', font=self.base_font)
        entry.place(x=x, y=y, width=width, height=height)
        entry.insert(0, value)
        entry.focus_force()

        def save_edit(evt):
            new_value = entry.get()
            entry.destroy()
            if new_value != value: 
                self.save_property_edit(item_id, new_value)
        
        entry.bind("<Return>", save_edit)
        entry.bind("<FocusOut>", save_edit)
        entry.bind("<Escape>", lambda e: entry.destroy())

    def populate_project_tree(self, element, parent_node_id):
        element_name = element.Name or f"ID: {element.GlobalId[:8]}..."
        node_id = self.project_tree.insert(parent_node_id, 'end', text=f"{element.is_a()} | {element_name}", open=True)
        self.tree_to_element_map[node_id] = element
        self.node_parent_map[node_id] = parent_node_id

        if hasattr(element, 'IsDecomposedBy'):
            for rel in element.IsDecomposedBy:
                for child in rel.RelatedObjects: self.populate_project_tree(child, node_id)
        
        if hasattr(element, 'ContainsElements'):
            for rel in element.ContainsElements:
                for child_element in rel.RelatedElements:
                    child_name = child_element.Name or f"ID: {child_element.GlobalId[:8]}..."
                    child_node_id = self.project_tree.insert(node_id, 'end', text=f"{child_element.is_a()} | {child_name}")
                    self.tree_to_element_map[child_node_id] = child_element
                    self.node_parent_map[child_node_id] = node_id
                    
    def on_tree_select(self, event=None):
        selected_items = self.project_tree.selection()
        if not selected_items:
            self.view_selected_button.config(state="disabled")
            return
        
        self.view_selected_button.config(state="normal")
        element = self.tree_to_element_map.get(selected_items[0])
        
        if element:
            self.display_key_properties(element)
            self.display_element_properties(element)
            
            if self.vedo_plotter and self.vedo_plotter.window:
                if self.transparency_var.get():
                    self.toggle_transparency_mode()
                else:
                    self.highlight_element_in_full_model(element)
            
            self.notebook.select(0)
            
    def display_key_properties(self, element):
        for label in self.key_props_labels.values(): label.config(text="N/A")
        for label in self.marks_labels.values(): label.config(text="N/A")
        self.description_text_label.config(text="N/A")

        all_psets = ifcopenshell.util.element.get_psets(element)
        
        def get_prop_val_str(names, unit=""):
            val = self._get_property_from_psets(all_psets, names if isinstance(names, list) else [names])
            if val is None: return "N/A"
            if isinstance(val, float): return f"{val:.2f}{unit}"
            return f"{str(val)}{unit}"

        self.key_props_labels["CATEGORÍA"].config(text=element.is_a())
        self.key_props_labels["FAMILIA"].config(text=get_prop_val_str("Familia"))
        self.key_props_labels["TIPO"].config(text=get_prop_val_str("Tipo"))
        container = ifcopenshell.util.element.get_container(element)
        level_name = container.Name if container and container.is_a("IfcBuildingStorey") else "N/A"
        self.key_props_labels["NIVEL"].config(text=level_name)
        self.key_props_labels["ÁREA"].config(text=get_prop_val_str(self.AREA_PARAM_NAMES, " m²"))
        self.key_props_labels["VOLUMEN"].config(text=get_prop_val_str(self.VOLUME_PARAM_NAMES, " m³"))

        self.marks_labels["CÓDIGO DE MONTAJE"].config(text=get_prop_val_str(self.ASSEMBLY_CODE_PARAM_NAMES))
        self.marks_labels["NOTA CLAVE"].config(text=get_prop_val_str(self.KEYNOTE_PARAM_NAMES))
        self.marks_labels["MARCA DE TIPO"].config(text=get_prop_val_str(self.TYPE_MARK_PARAM_NAMES))
        self.marks_labels["MARCA"].config(text=get_prop_val_str(self.MARK_PARAM_NAMES))
        
        type_element = ifcopenshell.util.element.get_type(element)
        type_name = type_element.Name if type_element else "N/A"
        self.marks_labels["ID DE TIPO"].config(text=type_name)
        self.marks_labels["GLOBALID"].config(text=element.GlobalId)
        
        description_value = get_prop_val_str(self.DESCRIPTION_PARAM_NAMES)
        self.description_text_label.config(text=description_value)
        
    def display_element_properties(self, element):
        self.properties_tree.delete(*self.properties_tree.get_children())
        
        info_id = self.properties_tree.insert('', 'end', text="INFORMACIÓN GENERAL", open=True, tags=('group_header',))
        self.properties_tree.insert(info_id, 'end', text="GlobalId", values=(str(element.GlobalId),))
        self.properties_tree.insert(info_id, 'end', text="Tipo", values=(str(element.is_a()),))
        if hasattr(element, 'Name'): self.properties_tree.insert(info_id, 'end', text="Nombre", values=(str(element.Name or ''),))
        if hasattr(element, 'Description'): self.properties_tree.insert(info_id, 'end', text="Description", values=(str(element.Description or ''),))
        
        all_sets = ifcopenshell.util.element.get_psets(element)
        
        def populate_set_tree(parent_node, title, data):
            if data:
                node_id = self.properties_tree.insert(parent_node, 'end', text=title.upper(), open=True, tags=('group_header',))
                for set_name, properties in data.items():
                    set_id = self.properties_tree.insert(node_id, 'end', text=set_name, open=True)
                    for prop_name, prop_value in properties.items():
                        if prop_name.lower() != 'id': self.properties_tree.insert(set_id, 'end', text=str(prop_name), values=(str(prop_value),))
        
        psets_data = {k: v for k, v in all_sets.items() if not self.ifc_file.by_id(v['id']).is_a('IfcQuantitySet')}
        qsets_data = {k: v for k, v in all_sets.items() if self.ifc_file.by_id(v['id']).is_a('IfcQuantitySet')}
        
        populate_set_tree('', "Propiedades (Property Sets)", psets_data)
        populate_set_tree('', "Cantidades (Quantity Sets)", qsets_data)
        
    def populate_summary_tab(self):
        for widget in self.summary_tab.winfo_children():
            widget.destroy()
        if not self.ifc_file: return
        
        summary_data = {}
        for element in self.ifc_file.by_type("IfcProduct"):
            container = ifcopenshell.util.element.get_container(element)
            level_name = container.Name if container and container.is_a("IfcBuildingStorey") else "Sin Nivel Asignado"
            if level_name not in summary_data: summary_data[level_name] = {}
            type_name = self.TRANSLATION_MAP.get(element.is_a(), element.is_a())
            summary_data[level_name][type_name] = summary_data[level_name].get(type_name, 0) + 1
            
        for level, types in sorted(summary_data.items()):
            total_elements = sum(types.values())
            level_frame = ttk.Labelframe(self.summary_tab, text=f"{level.upper()} ({total_elements} Elementos)", bootstyle="info", padding=15)
            level_frame.pack(fill='x', padx=10, pady=5)
            
            row = 0
            for type_name, count in sorted(types.items(), key=lambda item: item[1], reverse=True):
                ttk.Label(level_frame, text=type_name, font=self.base_font).grid(row=row, column=0, sticky='w', pady=2)
                ttk.Label(level_frame, text=count, font=self.bold_font).grid(row=row, column=1, sticky='e', padx=10, pady=2)
                row += 1
            level_frame.grid_columnconfigure(0, weight=1)
            
    
    def update_cost_tab(self):
        if not self.ifc_file: return

        for widget in self.cost_bar_chart_frame.winfo_children(): widget.destroy()

        elements = self.ifc_file.by_type("IfcProduct")
        total_cost, cost_by_type = self._calculate_costs(elements)
        
        self._update_cost_metrics_labels(total_cost, self.total_project_area, self.total_project_volume)
        self._update_timestamp_label()
        
        if cost_by_type:
            sorted_cost = dict(sorted(cost_by_type.items(), key=lambda item: item[1]))
            self._draw_cost_bar_chart(sorted_cost, total_cost)
        else:
            ttk.Label(self.cost_bar_chart_frame, text="No se encontraron parámetros de 'Costo' en el modelo.", justify=tk.CENTER).pack(pady=20)
            
    def update_spaces_tab(self):
        if not self.ifc_file: return
        
        self.spaces_tree.delete(*self.spaces_tree.get_children())
        
        spaces = self.ifc_file.by_type("IfcSpace")
        self._populate_spaces_table(spaces)
        
    def _populate_spaces_table(self, spaces):
        total_area, total_volume = 0.0, 0.0
        if not spaces: 
            self.total_project_area = 0.0
            self.total_project_volume = 0.0
            return

        spaces_info = []
        for space in spaces:
            psets = ifcopenshell.util.element.get_psets(space, psets_only=False)
            
            area_val = self._get_property_from_psets(psets, self.AREA_PARAM_NAMES) or 0.0
            volume_val = self._get_property_from_psets(psets, self.VOLUME_PARAM_NAMES) or 0.0
            
            space_number = self._get_property_from_psets(psets, self.SPACE_NUMBER_PARAM_NAMES) or space.Name or "N/A"
            space_name = self._get_property_from_psets(psets, self.SPACE_NAME_PARAM_NAMES) or space.LongName or "N/A"
            clear_height = self._get_property_from_psets(psets, self.CLEAR_HEIGHT_PARAM_NAMES)
            
            phase = self._get_property_from_psets(psets, self.PHASE_PARAM_NAMES, pset_name="PROCESO POR FASES") or "N/A"
            
            level_name = self._get_property_from_psets(psets, self.LEVEL_PARAM_NAMES)
            if not level_name:
                container = ifcopenshell.util.element.get_container(space)
                level_name = container.Name if container and container.is_a("IfcBuildingStorey") else "N/A"

            total_area += float(area_val)
            total_volume += float(volume_val)
            
            spaces_info.append({
                'number': str(space_number),
                'name': str(space_name),
                'level': str(level_name),
                'height': f"{clear_height:.2f}" if clear_height is not None else "N/A",
                'phase': str(phase),
                'area': f"{area_val:.2f}" if area_val > 0 else "N/A",
                'volume': f"{volume_val:.2f}" if volume_val > 0 else "N/A"
            })

        spaces_info.sort(key=lambda s: (s['level'], s['number']))

        last_level = None
        for info in spaces_info:
            if last_level is not None and info['level'] != last_level:
                self.spaces_tree.insert('', 'end', values=('', '', '', '', '', '', ''), tags=('spacer_row',))
            
            self.spaces_tree.insert('', 'end', text=info['number'], values=(
                info['name'], info['level'], info['height'], 
                info['phase'], info['area'], info['volume']
            ))
            last_level = info['level']

        self.spaces_tree.config(height=len(spaces_info) + (len(set(s['level'] for s in spaces_info)) -1) + 2)
        self.spaces_tree.insert('', 'end', text='TOTAL', 
                                 values=('', '', '', '', '', f"{total_area:.2f}", f"{total_volume:.2f}"), 
                                 tags=('total_row',))
            
        self.total_project_area = total_area
        self.total_project_volume = total_volume
        
    def _draw_spaces_by_level_donut(self, spaces_data):
        data = dict(sorted(spaces_data.items(), key=lambda item: item[1], reverse=True))
        labels = data.keys()
        sizes = data.values()
        
        bg_color, text_color = self.style.colors.bg, self.style.colors.fg
        colors = plt.cm.plasma(np.linspace(0.2, 0.8, len(labels)))

        fig = Figure(figsize=(10, 6), dpi=100, facecolor=bg_color)
        self.figures['spaces_donut'] = fig
        ax = fig.add_subplot(111)
        ax.set_facecolor(bg_color)

        def func(pct, allvals):
            absolute = int(round(pct/100.*np.sum(allvals)))
            return f"{absolute} habs.\n({pct:.1f}%)"

        wedges, texts, autotexts = ax.pie(sizes, 
                                          autopct=lambda pct: func(pct, sizes),
                                          startangle=90, 
                                          colors=colors,
                                          pctdistance=0.85,
                                          wedgeprops=dict(width=0.4, edgecolor=bg_color))
        
        plt.setp(autotexts, size=9, weight="bold", color="white")
        ax.axis('equal')

        legend_labels = [f'{l} ({s})' for l, s in zip(labels, sizes)]
        leg = ax.legend(wedges, legend_labels,
                          title="Niveles",
                          loc="center left",
                          bbox_to_anchor=(1, 0, 0.5, 1),
                          labelcolor=text_color,
                          frameon=False)
        leg.get_title().set_color(text_color)
        
        fig.tight_layout()
        canvas = FigureCanvasTkAgg(fig, master=self.spaces_donut_chart_frame)
        canvas.draw()
        canvas.get_tk_widget().pack(fill='both', expand=True, pady=5)


    def _calculate_costs(self, elements):
        total_cost = 0.0
        cost_by_type = Counter()
        
        for element in elements:
            psets = ifcopenshell.util.element.get_psets(element)
            cost_value = self._get_property_from_psets(psets, self.COST_PARAM_NAMES)
            
            if cost_value:
                try:
                    numeric_cost = float(cost_value)
                    if numeric_cost > 0:
                        total_cost += numeric_cost
                        element_type = self.TRANSLATION_MAP.get(element.is_a(), element.is_a())
                        cost_by_type[element_type] += numeric_cost
                except (ValueError, TypeError):
                    continue
        
        return total_cost, cost_by_type
        
    def _update_cost_metrics_labels(self, total_cost, total_area, total_volume):
        cost_per_sqm = total_cost / total_area if total_area > 0 else 0.0
        levels = self.ifc_file.by_type("IfcBuildingStorey")
        spaces = self.ifc_file.by_type("IfcSpace")

        self.metric_labels["cost"].config(text=locale.currency(total_cost, grouping=True))
        self.metric_labels["cost_sqm"].config(text=f"{locale.currency(cost_per_sqm, grouping=True)} / m²")
        self.metric_labels["area"].config(text=f"{total_area:.2f} m²")
        self.metric_labels["volume"].config(text=f"{total_volume:.2f} m³")
        self.metric_labels["levels"].config(text=f"{len(levels)}")
        self.metric_labels["spaces"].config(text=f"{len(spaces)}")

    def _update_timestamp_label(self):
        now = datetime.now()
        time_str = now.strftime("%d/%m/%Y a las %H:%M:%S")
        self.timestamp_label.config(text=f"Estimación de coste del modelo a fecha de {time_str}")

    def _draw_cost_bar_chart(self, data, total_cost):
        bg_color = self.style.colors.bg
        text_color = self.style.colors.fg
        primary_color = self.style.colors.primary

        labels = list(data.keys())
        values = list(data.values())
        
        fig = Figure(figsize=(10, len(labels) * 0.5), dpi=100, facecolor=bg_color)
        self.figures['cost'] = fig
        ax = fig.add_subplot(111)
        ax.set_facecolor(bg_color)
        
        colors = plt.cm.viridis(np.linspace(0.4, 0.9, len(labels)))
        y_pos = np.arange(len(labels))
        ax.barh(y_pos, values, color=colors, align='center', height=0.6)

        ax.spines['top'].set_visible(False); ax.spines['right'].set_visible(False)
        ax.spines['left'].set_visible(False); ax.spines['bottom'].set_color(text_color)
        ax.set_yticks([])
        ax.tick_params(axis='x', colors=text_color)
        ax.xaxis.set_major_formatter(lambda x, pos: locale.currency(x, symbol=True, grouping=True))

        for i, (label, value) in enumerate(data.items()):
            percent = (value / total_cost) * 100 if total_cost > 0 else 0
            category_text = f" {label.upper()}"
            value_text = f" {locale.currency(value, grouping=True)} ({percent:.1f}%)"
            ax.text(-0.01 * max(values), y_pos[i], category_text, ha='right', va='center', color=primary_color, weight='bold', fontsize=9)
            ax.text(0.01 * max(values), y_pos[i], value_text, ha='left', va='center', color=text_color, weight='normal', fontsize=9)

        fig.tight_layout(pad=1.5)
        
        canvas = FigureCanvasTkAgg(fig, master=self.cost_bar_chart_frame)
        canvas.draw()
        canvas.get_tk_widget().pack(fill='both', expand=True, pady=5)

    def export_summary_to_csv(self):
        if not self.ifc_file:
            Messagebox.show_warning("No hay datos para exportar.", "Sin Datos")
            return
            
        filepath = filedialog.asksaveasfilename(defaultextension=".csv", filetypes=[("Archivos CSV", "*.csv")], title="Guardar Resumen como CSV")
        if not filepath: return
        
        try:
            with open(filepath, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow(["Nivel", "Tipo de Elemento", "Cantidad"])
                
                summary_data = {}
                for element in self.ifc_file.by_type("IfcProduct"):
                    container = ifcopenshell.util.element.get_container(element)
                    level_name = container.Name if container and container.is_a("IfcBuildingStorey") else "Sin Nivel Asignado"
                    if level_name not in summary_data: summary_data[level_name] = {}
                    type_name = self.TRANSLATION_MAP.get(element.is_a(), element.is_a())
                    summary_data[level_name][type_name] = summary_data[level_name].get(type_name, 0) + 1
                    
                for level, types in sorted(summary_data.items()):
                    for type_name, count in sorted(types.items()):
                        writer.writerow([level, type_name, count])

            Messagebox.ok(f"Resumen guardado en {filepath}", "Exportación Completa")
        except Exception as e:
            Messagebox.show_error(f"No se pudo guardar el archivo.\nError: {e}", "Error de Exportación")

    def export_properties_to_csv(self):
        if not self.properties_tree.get_children():
            Messagebox.show_warning("No hay propiedades seleccionadas para exportar.", "Sin Datos")
            return
        
        filepath = filedialog.asksaveasfilename(defaultextension=".csv", filetypes=[("Archivos CSV", "*.csv")], title="Guardar Propiedades como CSV")
        if not filepath: return
        
        try:
            with open(filepath, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow(["Grupo", "Propiedad", "Valor"])
                for group_id in self.properties_tree.get_children():
                    group_name = self.properties_tree.item(group_id, "text")
                    for pset_id in self.properties_tree.get_children(group_id):
                        pset_name = self.properties_tree.item(pset_id, "text")
                        for prop_id in self.properties_tree.get_children(pset_id):
                            prop_name = self.properties_tree.item(prop_id, "text")
                            prop_value = self.properties_tree.item(prop_id, "values")[0]
                            writer.writerow([f"{group_name} | {pset_name}", prop_name, prop_value])
            Messagebox.ok(f"Propiedades guardadas en {filepath}", "Exportación Completa")
        except Exception as e:
            Messagebox.show_error(f"No se pudo guardar el archivo.\nError: {e}", "Error de Exportación")

    def _mark_as_modified(self):
        if not self.is_modified:
            self.is_modified = True
            self.save_button.config(state="normal")
            self.file_menu.entryconfig("Guardar como...", state="normal")

    def save_property_edit(self, item_id, new_value):
        selected_project_items = self.project_tree.selection()
        if not selected_project_items: return
        
        element = self.tree_to_element_map.get(selected_project_items[0])
        if not element: return
        
        prop_name = self.properties_tree.item(item_id, 'text')
        parent_id = self.properties_tree.parent(item_id)
        parent_text = self.properties_tree.item(parent_id, 'text')

        if parent_text == "INFORMACIÓN GENERAL":
            self._edit_direct_attribute(element, item_id, prop_name, new_value)
        else:
            grandparent_id = self.properties_tree.parent(parent_id)
            if grandparent_id and self.properties_tree.item(grandparent_id, 'text') == "PROPIEDADES (PROPERTY SETS)":
                self._edit_pset_property(element, item_id, parent_text, prop_name, new_value)

    def _edit_direct_attribute(self, element, item_id, prop_name, new_value):
        attribute_map = {"Nombre": "Name", "Description": "Description"}
        attribute_name = attribute_map.get(prop_name)
        if not attribute_name: return
        
        try:
            ifcopenshell.api.run("entity.edit_attributes", self.ifc_file, product=element, attributes={attribute_name: new_value})
            self.properties_tree.set(item_id, 'Value', new_value)
            self.properties_tree.item(item_id, tags=('edited',))
            self._mark_as_modified()
            
            if attribute_name == "Name":
                selected_item = self.project_tree.selection()[0]
                self.project_tree.item(selected_item, text=f"{element.is_a()} | {new_value}")
        except Exception as e:
            Messagebox.show_error(f"No se pudo cambiar el valor.\n\nError: {e}", "Error al Editar Atributo")

    def _edit_pset_property(self, element, item_id, pset_name, prop_name, new_value):
        try:
            psets = ifcopenshell.util.element.get_psets(element)
            target_pset_data = psets.get(pset_name)
            if not target_pset_data: return
            
            original_prop = target_pset_data.get(prop_name)
            if isinstance(original_prop, float): new_value = float(new_value.replace(',', '.'))
            elif isinstance(original_prop, int): new_value = int(new_value)
            
            pset_id = target_pset_data.get('id')
            if not pset_id: return
            pset_entity = self.ifc_file.by_id(pset_id)

            ifcopenshell.api.run("pset.edit_pset", self.ifc_file, pset=pset_entity, properties={prop_name: new_value})
            
            self.properties_tree.set(item_id, 'Value', new_value)
            self.properties_tree.item(item_id, tags=('edited',))
            self._mark_as_modified()
        except Exception as e:
            Messagebox.show_error(f"No se pudo cambiar el valor.\nVerifica que el tipo de dato sea correcto (ej. número, texto).\n\nError: {e}", "Error al Editar Propiedad")

    def save_modified_ifc(self):
        if not self.ifc_file:
            Messagebox.show_warning("No hay ningún archivo IFC cargado.", "Sin Archivo")
            return
        if not self.is_modified:
            Messagebox.show_info("No se ha realizado ninguna modificación.", "Sin Cambios")
            return
            
        base, ext = os.path.splitext(os.path.basename(self.ifc_filepath))
        suggested_filename = f"{base}_modificado.ifc"
        
        filepath = filedialog.asksaveasfilename(title="Guardar Archivo IFC Modificado", initialfile=suggested_filename, defaultextension=".ifc", filetypes=(("Archivos IFC", "*.ifc"),("Todos", "*.*")))
        if not filepath: return
        
        try:
            self.update_status("Guardando archivo modificado...", 0)
            self.ifc_file.write(filepath)
            self.update_status("¡Archivo guardado con éxito!", 100)
            Messagebox.ok(f"El archivo modificado ha sido guardado en:\n{filepath}", "Guardado Completo")
            self.is_modified = False
            self.save_button.config(state="disabled")
            self.file_menu.entryconfig("Guardar como...", state="disabled")
            self.root.after(2000, lambda: self.update_status("Listo.", 0))
        except Exception:
            self.update_status("Error al guardar.", 0)
            handle_error(traceback.format_exc(), self.root)

    def collect_searchable_parameters(self):
        if not self.ifc_file: return
        param_set = set()
        for product in self.ifc_file.by_type('IfcProduct'):
            psets = ifcopenshell.util.element.get_psets(product)
            for properties in psets.values():
                for prop_name in properties.keys():
                    if prop_name.lower() != 'id': param_set.add(prop_name)
                    
        base_params = ["Nombre de Elemento", "ID de Elemento", "Tipo IFC (ej. IfcDoor)"]
        self.searchable_parameters = base_params + sorted(list(param_set))

    def open_advanced_search(self):
        if not self.ifc_file:
            Messagebox.show_warning("Carga primero un archivo IFC.", "Sin Archivo")
            return
            
        win = tk.Toplevel(self.root)
        win.title("Búsqueda Avanzada")
        win.transient(self.root)
        win.geometry("400x150")
        
        ttk.Label(win, text="Buscar por:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
        combo = ttk.Combobox(win, values=self.searchable_parameters, width=40, state="readonly")
        combo.grid(row=0, column=1, padx=5, pady=5)
        if self.searchable_parameters: combo.current(0)
        
        ttk.Label(win, text="Valor:").grid(row=1, column=0, padx=5, pady=5, sticky="w")
        entry = ttk.Entry(win, width=42)
        entry.grid(row=1, column=1, padx=5, pady=5)
        entry.focus_set()

        def perform_search():
            search_by = combo.get()
            query = entry.get().lower().strip()
            if not query: return
            
            self._restore_full_tree_view()
            
            items_to_show = set()
            for item_id, element in self.tree_to_element_map.items():
                match = False
                if search_by == "Nombre de Elemento" and element.Name and query in element.Name.lower():
                    match = True
                elif search_by == "ID de Elemento" and query in element.GlobalId.lower():
                    match = True
                elif search_by == "Tipo IFC (ej. IfcDoor)" and query in element.is_a().lower():
                    match = True
                else:
                    psets = ifcopenshell.util.element.get_psets(element)
                    prop_value = self._get_property_from_psets(psets, [search_by])
                    if prop_value is not None and query in str(prop_value).lower():
                        match = True

                if match:
                    items_to_show.add(item_id)
                    parent = self.node_parent_map.get(item_id)
                    while parent:
                        items_to_show.add(parent)
                        parent = self.node_parent_map.get(parent)
            
            items_to_hide = set(self.tree_to_element_map.keys()) - items_to_show
            for item_id in items_to_hide:
                self.project_tree.detach(item_id)
            win.destroy()
        
        ttk.Button(win, text="Buscar", command=perform_search).grid(row=2, column=1, padx=5, pady=10, sticky="e")
        win.bind('<Return>', lambda e: perform_search())

    def _restore_full_tree_view(self):
        for node_id, parent_id in self.node_parent_map.items():
            if self.project_tree.exists(node_id):
                self.project_tree.move(node_id, parent_id, 'end')

    def simple_search(self, event=None):
        query = self.search_entry.get().lower().strip()
        
        self._restore_full_tree_view()
        if not query:
            return

        items_to_show = set()
        for item_id in self.tree_to_element_map:
            if query in self.project_tree.item(item_id, 'text').lower():
                items_to_show.add(item_id)
                parent = self.node_parent_map.get(item_id)
                while parent:
                    items_to_show.add(parent)
                    parent = self.node_parent_map.get(parent)

        items_to_hide = set(self.tree_to_element_map.keys()) - items_to_show
        for item_id in items_to_hide:
            self.project_tree.detach(item_id)

    def clear_simple_search(self, event=None):
        self.search_entry.delete(0, tk.END)
        self._restore_full_tree_view()

    def open_export_dialog(self, mode='model'):
        if not self.ifc_file:
            Messagebox.show_warning("No hay ningún archivo IFC cargado.", "Sin Archivo")
            return

        if mode == 'element' and not self.project_tree.selection():
            Messagebox.show_warning("Selecciona un elemento", "Por favor, selecciona un elemento del árbol para exportar.")
            return

        suggested_filename = ""
        if mode == 'model':
            base, _ = os.path.splitext(os.path.basename(self.ifc_filepath))
            suggested_filename = f"Informe_Modelo_{base}.xlsx"
        else:
            selected_item = self.project_tree.selection()[0]
            element = self.tree_to_element_map.get(selected_item)
            element_name = element.Name.replace(" ", "_") if element.Name else element.GlobalId
            suggested_filename = f"Informe_Elemento_{element_name[:20]}.xlsx"

        filepath = filedialog.asksaveasfilename(
            title="Guardar Informe Excel",
            initialfile=suggested_filename,
            defaultextension=".xlsx",
            filetypes=[("Archivos Excel", "*.xlsx")]
        )
        if not filepath:
            return

        threading.Thread(target=self._export_to_excel, args=(filepath, mode), daemon=True).start()

    def _export_to_excel(self, filepath, mode):
        self.update_status(f"Exportando a Excel...", 0)
        try:
            if mode == 'model':
                self.export_model_report(filepath)
            else:
                self.export_element_report(filepath)
            
            self.update_status(f"Informe guardado en {os.path.basename(filepath)}", 100)
            self.root.after(0, lambda: Messagebox.ok("Exportación a Excel completada con éxito.", "Exportación Completa"))
        except Exception:
            error_info = traceback.format_exc()
            self.root.after(0, lambda: handle_error(f"Error durante la exportación a Excel:\n{error_info}", self.root))
        finally:
            self.root.after(2000, lambda: self.update_status("Listo.", 0))

    def export_model_report(self, filepath):
        wb = openpyxl.Workbook()
        
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        self._populate_summary_sheet(wb.create_sheet("Resumen"))
        self._populate_cost_sheet(wb.create_sheet("Coste"))
        self._populate_spaces_excel_sheet(wb.create_sheet("Espacios"))
        self._populate_analysis_sheet(wb.create_sheet("Análisis"))
        
        wb.save(filepath)

    def export_element_report(self, filepath):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Informe de Elemento"

        selected_item = self.project_tree.selection()[0]
        element = self.tree_to_element_map.get(selected_item)
        if not element: return

        title_font = Font(name='Calibri', size=16, bold=True, color="1F497D")
        header_font = Font(name='Calibri', size=12, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        
        row_cursor = 3
        
        def write_block(title, labels_dict, start_row):
            ws[f'A{start_row}'] = title
            ws[f'A{start_row}'].font = Font(bold=True, size=12)
            ws.merge_cells(f'A{start_row}:B{start_row}')
            r = start_row + 1
            for key, label_widget in labels_dict.items():
                ws[f'A{r}'] = key
                ws[f'A{r}'].font = self.bold_font
                ws[f'B{r}'] = label_widget.cget("text")
                r += 1
            return r + 1

        row_cursor = write_block("Propiedades Clave", self.key_props_labels, row_cursor)
        row_cursor = write_block("Marcas", self.marks_labels, row_cursor)
        
        ws[f'A{row_cursor}'] = "Descripción"
        ws[f'A{row_cursor}'].font = Font(bold=True, size=12)
        ws.merge_cells(f'A{row_cursor}:D{row_cursor}')
        desc_cell = ws[f'A{row_cursor + 1}']
        desc_cell.value = self.description_text_label.cget("text")
        desc_cell.alignment = Alignment(wrap_text=True, vertical='top')
        ws.merge_cells(f'A{row_cursor + 1}:D{row_cursor + 4}')
        row_cursor += 6
        
        ws[f'A{row_cursor}'] = "Listado Completo de Propiedades"
        ws[f'A{row_cursor}'].font = title_font
        ws.merge_cells(f'A{row_cursor}:C{row_cursor}')
        row_cursor += 1
        
        headers = ["Grupo / Pset", "Propiedad", "Valor"]
        for i, header in enumerate(headers):
            cell = ws.cell(row=row_cursor, column=i+1, value=header)
            cell.font = header_font
            cell.fill = header_fill
        row_cursor += 1
        
        for group_id in self.properties_tree.get_children():
            group_name = self.properties_tree.item(group_id, "text")
            for pset_id in self.properties_tree.get_children(group_id):
                pset_name = self.properties_tree.item(pset_id, "text")
                ws.cell(row=row_cursor, column=1, value=f"{group_name} | {pset_name}").font = self.bold_font
                row_cursor +=1
                for prop_id in self.properties_tree.get_children(pset_id):
                    prop_name = self.properties_tree.item(prop_id, "text")
                    prop_value = self.properties_tree.item(prop_id, "values")[0]
                    ws.cell(row=row_cursor, column=2, value=prop_name)
                    ws.cell(row=row_cursor, column=3, value=prop_value)
                    row_cursor += 1
            row_cursor += 1

        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 50

        wb.save(filepath)

    def _prepare_figure_for_export(self, fig):
        fig.set_facecolor('white')
        for ax in fig.get_axes():
            ax.set_facecolor('white')
            
            for spine in ax.spines.values():
                spine.set_color('black')
            ax.tick_params(axis='x', colors='black')
            ax.tick_params(axis='y', colors='black')
            
            ax.yaxis.label.set_color('black')
            ax.xaxis.label.set_color('black')
            ax.title.set_color('black')
            
            for text in ax.texts:
                if hasattr(text, 'set_color'):
                    text.set_color('black')

            if ax.get_legend():
                leg = ax.get_legend()
                leg.get_title().set_color('black')
                for text in leg.get_texts():
                    text.set_color('black')
            
            if any(isinstance(p, plt.matplotlib.patches.Wedge) for p in ax.patches):
                for child in ax.get_children():
                    if isinstance(child, plt.Annotation):
                        child.set_color('black')
                    if hasattr(child, 'arrow_patch') and child.arrow_patch:
                         child.arrow_patch.set_facecolor('black')

    def _populate_summary_sheet(self, ws):
        ws.title = "Resumen"
        row_cursor = 1
        title_font = Font(bold=True, size=14)
        
        summary_data = {}
        for element in self.ifc_file.by_type("IfcProduct"):
            container = ifcopenshell.util.element.get_container(element)
            level_name = container.Name if container and container.is_a("IfcBuildingStorey") else "Sin Nivel Asignado"
            if level_name not in summary_data: summary_data[level_name] = {}
            type_name = self.TRANSLATION_MAP.get(element.is_a(), element.is_a())
            summary_data[level_name][type_name] = summary_data[level_name].get(type_name, 0) + 1
        
        for level, types in sorted(summary_data.items()):
            total_elements = sum(types.values())
            
            ws.cell(row=row_cursor, column=1, value=f"{level.upper()}").font = title_font
            ws.cell(row=row_cursor, column=2, value=f"{total_elements} Elementos").font = title_font
            ws.merge_cells(start_row=row_cursor, start_column=2, end_row=row_cursor, end_column=3)
            row_cursor += 1
            
            for type_name, count in sorted(types.items(), key=lambda item: item[1], reverse=True):
                ws.cell(row=row_cursor, column=2, value=type_name)
                ws.cell(row=row_cursor, column=3, value=count)
                row_cursor += 1
            row_cursor += 1
            
        self._format_sheet(ws, format_headers=False)

    def _populate_cost_sheet(self, ws):
        ws.title = "Coste"
        row_cursor = 1
        title_font = Font(bold=True, size=14)
        header_font = Font(bold=True)

        ws.cell(row=row_cursor, column=1, value="Resumen Económico y Geométrico").font = title_font
        row_cursor += 1
        
        metrics = [
            ("Coste Total", self.metric_labels['cost'].cget("text")),
            ("Coste / m²", self.metric_labels['cost_sqm'].cget("text")),
            ("Superficie Total Útil", self.metric_labels['area'].cget("text")),
            ("Volumen Total", self.metric_labels['volume'].cget("text")),
            ("Nº de Habitaciones", self.metric_labels['spaces'].cget("text")),
            ("Nº de Niveles", self.metric_labels['levels'].cget("text")),
        ]
        for key, val in metrics:
            ws.cell(row=row_cursor, column=1, value=key).font = header_font
            ws.cell(row=row_cursor, column=2, value=val)
            row_cursor += 1

        self._format_sheet(ws, format_headers=False)

    def _populate_spaces_excel_sheet(self, ws):
        ws.title = "Espacios"
        row_cursor = 1
        title_font = Font(bold=True, size=14)
        header_font = Font(bold=True)

        ws.cell(row=row_cursor, column=1, value="Listado de Espacios").font = title_font
        row_cursor += 2
        
        headers = ['Número', 'Nombre', 'Nivel', 'Altura Libre (m)', 'Fase', 'Superficie (m²)', 'Volumen (m³)']
        for col_idx, header_title in enumerate(headers, 1):
            cell = ws.cell(row=row_cursor, column=col_idx, value=header_title)
            cell.font = header_font
        row_cursor += 1
        
        for item_id in self.spaces_tree.get_children():
            row_data = [self.properties_tree.item(item_id, 'text')] + list(self.properties_tree.item(item_id, 'values'))
            for col_idx, cell_value in enumerate(row_data, 1):
                 ws.cell(row=row_cursor, column=col_idx, value=cell_value)
            row_cursor += 1
            
        self._format_sheet(ws, format_headers=False)


    def _populate_analysis_sheet(self, ws):
        ws.title = "Análisis"

        charts_to_export = [
            ('summary', "RECUENTO TOTAL POR NIVEL"),
            ('elements_by_level', "ELEMENTOS POR NIVELES"),
            ('donut', "DISTRIBUCIÓN PORCENTUAL DE ELEMENTOS")
        ]

        row_cursor = 2

        for fig_key, title in charts_to_export:
            fig = self.figures.get(fig_key)
            if not fig:
                continue

            title_cell = f'A{row_cursor}'
            ws[title_cell] = title
            ws[title_cell].font = Font(bold=True, size=14)
            row_cursor += 1

            fig_copy = copy.deepcopy(fig)
            self._prepare_figure_for_export(fig_copy)

            img_data = io.BytesIO()
            fig_copy.savefig(img_data, format='png', bbox_inches='tight', dpi=200)
            img_data.seek(0)
            img = OpenpyxlImage(img_data)
            
            aspect_ratio = fig_copy.get_figheight() / fig_copy.get_figwidth()
            img.height = 300 
            img.width = img.height / aspect_ratio

            img.anchor = f'A{row_cursor}'
            ws.add_image(img)

            row_cursor += 16
            
        self._format_sheet(ws, format_headers=False)

    def _format_sheet(self, ws, format_headers=True):
        if format_headers:
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            
            for row in ws.iter_rows(min_row=1, max_row=1):
                for cell in row:
                    cell.font = header_font
                    cell.fill = header_fill

        for column_cells in ws.columns:
            try:
                max_len = 0
                for cell in column_cells:
                    if not isinstance(cell, openpyxl.cell.cell.MergedCell):
                        if cell.value:
                           max_len = max(max_len, len(str(cell.value)))
                
                ws.column_dimensions[column_cells[0].column_letter].width = max_len + 2

            except Exception:
                pass

def handle_error(error_info_str, root_window=None):
    log_message = f"La aplicación ha encontrado un error irrecuperable.\nSe intentará guardar un informe de error en el escritorio."
    
    if root_window:
        Messagebox.show_error(log_message, "Error Crítico", parent=root_window)
    else:
        temp_root = tk.Tk()
        temp_root.withdraw()
        Messagebox.show_error(log_message, "Error Crítico")
        temp_root.destroy()
        
    try:
        desktop_path = os.path.join(os.path.join(os.environ.get('USERPROFILE', os.environ.get('HOME', '')), 'Desktop'))
        log_file_path = os.path.join(desktop_path, 'visor_ifc_error_log.txt')
        with open(log_file_path, 'w', encoding='utf-8') as f:
            f.write("--- ERROR CRÍTICO EN VISOR IFC ---\n\n")
            f.write(error_info_str)
    except Exception as log_e: 
        final_error_msg = f"No se pudo escribir el log de error en el escritorio: {log_e}\n\nError original:\n{error_info_str}"
        if root_window:
             Messagebox.showerror("Error Crítico de Log", final_error_msg, parent=root_window)
        else:
            tk.messagebox.showerror("Error Crítico de Log", final_error_msg)


def main():
    """Función principal para inicializar y ejecutar la aplicación."""
    missing_libs = []
    if not IFCOPENDHELL_AVAILABLE: missing_libs.append("ifcopenshell")
    if not VEDO_AVAILABLE: missing_libs.append("vedo")
    if not MATPLOTLIB_AVAILABLE: missing_libs.append("matplotlib")
    if not TTKBOOTSTRAP_AVAILABLE: missing_libs.append("ttkbootstrap")
    if not PANDAS_AVAILABLE: missing_libs.append("pandas")
    if not OPENPYXL_AVAILABLE: missing_libs.append("openpyxl")
    if not PILLOW_AVAILABLE: missing_libs.append("Pillow")

    if missing_libs:
        root = tk.Tk()
        root.withdraw()
        message = "Faltan librerías esenciales:\n\n" + "\n".join(missing_libs)
        message += "\n\nPor favor, instálalas para ejecutar la aplicación (ej: pip install matplotlib)."
        tk.messagebox.showerror("Faltan Librerías", message)
        return

    try:
        root = ttk.Window(themename="superhero")
        app = IFCViewerApp(root)
        root.mainloop()
    except Exception:
        error_info = traceback.format_exc()
        handle_error(error_info)

if __name__ == "__main__":
    main()
